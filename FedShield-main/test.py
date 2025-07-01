from openpyxl import load_workbook
import re

def add_words(wordsss):
    string=wordsss
    # Load the existing workbook
    workbook = load_workbook('4000.xlsx')

    # Select the active worksheet
    sheet = workbook.active
    words=re.sub(r"\s+", "", string)
    wordes= words.split(',')
    # Define the word to add
    for i in range(len(wordes)):
        # Find the first empty cell in the first column
        for row in range(1, sheet.max_row + 2):  # Adjusting range to check one row beyond the current max
            cell_value = sheet.cell(row=row, column=1).value
            if cell_value is None:
                sheet.cell(row=row, column=1, value=wordes[i])
                break

    # Save the workbook
    workbook.save('4000.xlsx')
