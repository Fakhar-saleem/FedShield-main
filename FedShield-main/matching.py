import hashlib
import os

import pandas as pd
import openpyxl as xl
from docx import Document

def calculte_excel_hash(file_path_hash):
    wb = xl.load_workbook(file_path_hash)
    sheet_name = wb.sheetnames
    data = pd.read_excel(file_path_hash, sheet_name=sheet_name[0])
    # Assuming the first row is to be hashed
    row_to_hash = data.iloc[0].to_string()

    # Calculate the hash (using SHA-256 in this example)
    hash_object = hashlib.sha256(row_to_hash.encode())
    hashed_row = hash_object.hexdigest()

    print("Hash of the first row:", hashed_row)
#--------------------------------------------------------------------------------
def calculate_txt_hash(file_path):
    # Initialize the hash object based on the specified algorithm

    hash_object = hashlib.sha256()

    # Open the file and read it in binary mode
    try:
        with open(file_path, 'rb') as file:
            # Read the file in chunks to avoid loading the entire file into memory
            while chunk := file.read(8192):
                hash_object.update(chunk)
    except FileNotFoundError:
        print("File not found.")
        return None

    # Get the hexadecimal representation of the hash
    file_hash = hash_object.hexdigest()
    return file_hash
#--------------------------------------------------------------------------------
def keywordMatching(wordFile):
    name=os.path.basename(wordFile)
    _, ext = os.path.splitext(name)
    print(name,ext)
    filepath="4000.xlsx"
    wb = xl.load_workbook(filepath)
    sheet_name = wb.sheetnames
    for index, value in enumerate(sheet_name):
        df = pd.read_excel(filepath, sheet_name=sheet_name[index])
        keyword_array = df['*DLP*'].tolist()
        if ext == ".docx":
            doc = Document(wordFile)
            for para in doc.paragraphs:
                words = para.text.split()
                for word in words:
                    i = 1
                    for x in keyword_array:
                        if len(x) == len(word):
                            if x.lower() == word.lower():
                                print("Word Matched. Index No. ", i + 1)
                                return True
                        i = i + 1
            return False
        elif ext==".txt":
            with open(wordFile, 'r') as file:
                for line in file:
                    words = line.split()
                    for word in words:
                        i = 1
                        for x in keyword_array:
                            if len(x) == len(word):
                                if x.lower() == word.lower():
                                    print(f"Word Matched= {word}. Index No. ", i+1)
                                    return True
                            i = i + 1
            return False
        elif ext==".xlsx":
            workbook = xl.load_workbook(wordFile)
            with open('output.txt', 'w') as txt_file:
                for sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    txt_file.write(f"--- {sheet_name} ---\n")
                    for row in sheet.iter_rows():
                        for cell in row:
                            txt_file.write(str(cell.value) + ' ')
                        txt_file.write('\n')

            with open('output.txt', 'r') as file:
                for line in file:
                    words = line.split()
                    for word in words:
                        i = 1
                        for x in keyword_array:
                            if len(x) == len(word):
                                if x.lower() == word.lower():
                                    print(f"Word Matched= {word}. Index No. ", i + 1)
                                    file.close()
                                    os.remove('output.txt')
                                    return True
                            i = i + 1
            return False



#---------------------------------------------------------------------------------
